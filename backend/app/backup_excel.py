"""Respaldo en Excel de cada salida/entrada, en paralelo a la base de datos.

Pensado para cuando en unas semanas la app corra en el cluster del trabajo: si
el cluster se cae, este archivo queda como historial de respaldo — se puede
abrir y seguir anotando movimientos a mano en él, igual que con la planilla
vieja, mientras la app no esté disponible.

Reglas de diseño:
- Solo se AGREGAN filas al final, nunca se reescribe ni se borra nada ya
  escrito — así, si alguien anotó algo a mano mientras la app estaba caída,
  cuando la app vuelva a andar no le pisa esas filas.
- Si por algún motivo no se puede escribir (por ejemplo, alguien tiene el
  archivo abierto en Excel justo en ese momento), no debe romper el préstamo o
  la devolución real: la base de datos sigue siendo la fuente de verdad: el
  respaldo es "mejor esfuerzo", no algo de lo que dependa el flujo principal.
"""
import datetime
import threading

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .paths import data_dir

BACKUP_PATH = data_dir() / "BacKup.xlsx"
HOJA = "Movimientos"
ENCABEZADO = [
    "Fecha", "Hora", "Movimiento", "Código", "Elemento",
    "Curso / Evento", "Responsable", "Hora a devolver", "Cargado por",
]

# Varias peticiones pueden llegar en paralelo (distintas terminales escaneando
# a la vez) — sin este lock, dos escrituras casi simultáneas podrían leer el
# mismo archivo, y la que guarda último pisaría la fila de la otra.
_lock = threading.Lock()


def _crear_archivo_nuevo():
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA
    ws.append(ENCABEZADO)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="2F5D50", end_color="2F5D50", fill_type="solid")
    anchos = [12, 10, 11, 20, 26, 16, 22, 15, 18]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    wb.save(BACKUP_PATH)


def registrar(
    tipo: str,  # "SALIDA" o "ENTRADA"
    elemento_codigo: str,
    elemento_nombre: str | None,
    curso_evento: str | None,
    responsable: str | None,
    hora_devolver: str | None = None,
    cargado_por: str | None = None,
    momento: datetime.datetime | None = None,
):
    momento = momento or datetime.datetime.now()
    try:
        with _lock:
            if not BACKUP_PATH.exists():
                _crear_archivo_nuevo()
            wb = load_workbook(BACKUP_PATH)
            ws = wb[HOJA] if HOJA in wb.sheetnames else wb.active
            ws.append([
                momento.strftime("%d/%m/%Y"),
                momento.strftime("%H:%M:%S"),
                tipo,
                elemento_codigo,
                elemento_nombre or "",
                curso_evento or "",
                responsable or "",
                hora_devolver or "",
                cargado_por or "",
            ])
            wb.save(BACKUP_PATH)
    except Exception:
        # "mejor esfuerzo": si el respaldo falla (archivo abierto, disco lleno,
        # etc.) no debe tumbar el préstamo/devolución real.
        pass
