import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from . import models


def generar_excel_compras(compras: list["models.Compra"]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    encabezado = ["Fecha de carga", "Cantidad", "Qué comprar", "Prioridad"]
    ws.append(encabezado)
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="2F5D50", end_color="2F5D50", fill_type="solid")

    for c in compras:
        ws.append([
            c.fecha_hora.strftime("%d/%m/%Y %H:%M"),
            c.cantidad,
            c.descripcion,
            c.prioridad,
        ])

    anchos = [18, 10, 55, 12]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
